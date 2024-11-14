import json
import os.path
from datetime import datetime
import freecurrencyapi
from openpyxl import Workbook
import matplotlib
from matplotlib import cm
matplotlib.use('Agg')
import matplotlib.pyplot as plt

def devaluation_rate(initial_value, end_value):
    percentage_of_change = ((end_value - initial_value) / initial_value) * 100
    return percentage_of_change


class App:

    def __init__(self, api_key: str):
        self.api = freecurrencyapi.Client(api_key)
        self.currencies = {}
        self.exchanges = {}
        self.historical = {}
        self.historial = {"fecha": datetime.now().strftime("%d/%m/%Y"), "comandos": {}}

    def save_command(self, command, result = ""):
        self.historial["comandos"][datetime.now().strftime("%H:%M:%S")] = (command, result)
        pass

    def load_data(self):
        currencies_path = 'data/currencies.json'

        if os.path.exists(currencies_path):
            with open(currencies_path, 'r') as f:
                data = json.load(f)
                self.currencies = data['data']

        else:
            if not os.path.exists('data'):
                os.mkdir('data')

            data = self.api.currencies()
            self.currencies = data['data']

            with open(currencies_path, 'w') as f:
                json.dump(data, f)
                f.close()

    def supported_currencies(self):

        resArray = []

        for code, currency in self.currencies.items():
            res = (f'{code}: [{currency['symbol_native']}] - {currency['name']}')
            print(res)
            resArray.append(res)
        
        return resArray

    def search_currency(self, text: str):
        
        resArray = []

        for code, currency in self.currencies.items():
            name: str = currency['name']

            if text.lower() in name.lower():
                res = (f'{code}: [{currency['symbol_native']}] - {currency['name']}')
                print(res)
                resArray.append(res)

        return resArray

    def get_exchange_rate(self, base_currency: str, target_currency: str) -> float:

        if base_currency not in self.exchanges:
            response = self.api.latest(base_currency)
            self.exchanges[base_currency] = response['data']

        return self.exchanges[base_currency][target_currency]

    def get_historical_exchange_rate(self, base_currency: str, target_currency: str, date: str) -> float:

        if datetime.now().strftime('%Y-%m-%d') == date:
            return self.get_exchange_rate(base_currency, target_currency)
        else:

            if date not in self.historical:
                self.historical[date] = {}

            if base_currency not in self.historical[date]:

                response = self.api.historical(date, base_currency)
                self.historical[date][base_currency] = response['data'][date]

            return self.historical[date][base_currency][target_currency]

    def convert(self, amount: float, base_currency: str, currencies: list[str]):

        resArray = []

        if base_currency not in self.currencies.keys():
            res = ('Codigo de moneda base invalido: ' + base_currency)
            print(res)
            return res

        for currency in currencies:

            if currency not in self.currencies.keys():
                res = ('Codigo de moneda objetivo invalido: ' + currency)
                print(res)
                return res

            result = self.get_exchange_rate(base_currency, currency) * amount

            res = (f'{self.currencies[currency]['name']} ({self.currencies[currency]['code']}) -> '
                f'{result:.2f} {self.currencies[currency]['symbol_native']}')

            resArray.append(res)
            
            print(res)

        return resArray

    def stats(self, base_currency: str, currencies: list[str], initial_date: str, end_date: str):

        date1 = datetime.strptime(initial_date, '%Y-%m-%d').date()
        date2 = datetime.strptime(end_date, '%Y-%m-%d').date()

        if date2 < date1:
            res ='end_date no puede ser menor que start_date'
            print(res)
            return res
        elif date1 == date2:
            res ='start_date y end_date no pueden ser iguales'
            print(res)
            return res
        elif date2 > datetime.now().date():
            res = 'end_date no puede ser mayor que el dia actual'
            print(res)
            return res

        if base_currency not in self.currencies.keys():
            res = 'Codigo de moneda base invalido: ' + base_currency
            print(res)
            return res

        resArray = []

        for currency in currencies:
            
            if currency not in self.currencies.keys():
                res = 'Codigo de moneda base invalido: ' + base_currency
                print(res)
                return res

            exchange1 = self.get_historical_exchange_rate(base_currency, currency, initial_date)
            exchange2 = self.get_historical_exchange_rate(base_currency, currency, end_date)

            result = devaluation_rate(exchange1, exchange2)

            res = (f'''{self.currencies[currency]['name']} ({self.currencies[currency]['code']}):
    Exchange rate at {initial_date}: {exchange1:.2f} {self.currencies[currency]['symbol_native']}
    Exchange rate at {end_date}: {exchange2:.2f} {self.currencies[currency]['symbol_native']}
    Devaluation: {result:+.2f} %''')
            
            print(res)

            resArray.append(res)

        return resArray

    def export_Historial(self):

        wb = Workbook()

        hoja = wb.active

        hoja.title = f'Historial {datetime.now().strftime("%d-%m-%Y")}'
        hoja['A1'] = "Hora"
        hoja['B1'] = "Comando"
        hoja['C1'] = "Resultado"

        i = 2

        for clave, valor in self.historial["comandos"].items():
            hoja[f'A{i}'] = clave
            hoja[f'B{i}'] = valor[0]
            if isinstance(valor[1], list):
                for element in valor[1]:
                    hoja[(f'C{i}')] = element
                    i += 1
            elif isinstance(valor[1], str):
                hoja[(f'C{i}')] = valor[1]
                i += 1
            else:
                print("Error")
            
            i += 1

        hojaConversiones = wb.create_sheet("Datos_Conv")

        hojaConversiones['A1'] = "Moneda base"
        hojaConversiones['B1'] = "Moneda de cambio"
        hojaConversiones['C1'] = "Tipo de cambio"

        i=2

        for clave, valor in self.exchanges.items():
            hojaConversiones[f'A{i}'] = clave
            for clave2, valor2 in valor.items():
                hojaConversiones[f'B{i}'] = clave2
                hojaConversiones[f'C{i}'] = valor2

                i += 1
            
            i += 1

        # print(self.historical)

        hojaHistorical = wb.create_sheet("Datos_Stats")

        hojaHistorical['A1'] = "Fecha"
        hojaHistorical['B1'] = "Moneda base"
        hojaHistorical['C1'] = "Moneda de cambio"
        hojaHistorical['D1'] = "Tipo de cambio"

        i=2

        for clave, valor in self.historical.items():
            hojaHistorical[f'A{i}'] = clave
            for clave2, valor2 in valor.items():
                hojaHistorical[f'B{i}'] = clave2
                for clave3, valor3 in valor2.items():
                    hojaHistorical[f'C{i}'] = clave3
                    hojaHistorical[f'D{i}'] = valor3

                    i += 1

                i += 1
                
            i += 1

        wb.save("Historial.xlsx")

    # Seccion de graficas

    def graph_conv(self, base_currency, currencys):

        if base_currency not in self.currencies.keys():
            res = ('Codigo de moneda base invalido: ' + base_currency)
            print(res)
            return res

        values = []

        for element in currencys:
            if element not in self.currencies.keys():
                res = ('Codigo de moneda base invalido: ' + element)
                print(res)
                return res
        
            res = self.get_exchange_rate(base_currency, element)
            values.append(res)

        # Crear el gráfico de barras
        plt.figure(figsize=(10, 8))
        fig = plt.figure()
        ax = fig.add_subplot(111)
        ax.set_facecolor('thistle')
        colormap = cm.viridis
        plt.bar(currencys, values,  color=colormap(range(len(values))), edgecolor='black', linewidth=1.3, hatch='//')
        plt.grid(True, color='black', linestyle='--', linewidth=0.5)

        # Añadir etiquetas y título
        plt.xlabel('Moneda')
        plt.ylabel('Valor')
        plt.title(f'Tasas de Cambio {base_currency}')

        # Rotar las etiquetas del eje x para que no se solapen
        plt.xticks(rotation=90)

        # Mostrar el gráfico
        plt.tight_layout()

        nombre = './graficas/convercion.png'

        plt.savefig(nombre)

        print(f"Grafico creado en {nombre}")

    def get_month_exchange(self, mes, año, base_currency, currency):

        res = []

        # Ajustamos la lógica para la fecha de la petición
        for i in range(1, mes):

            if i < 10:
                fecha_peticion = f'{año}-0{i}-15'
            else:
                fecha_peticion = f'{año}-{i}-15'

            # Hacer la petición para el tipo de cambio histórico
            res.append(self.get_historical_exchange_rate(base_currency, currency, fecha_peticion))

        # Obtener el tipo de cambio actual
        res.append(self.get_exchange_rate(base_currency, currency))

        return res

    def graph_historial(self, base_currency, currency):

        if base_currency not in self.currencies.keys():
            res = ('Codigo de moneda base invalido: ' + base_currency)
            print(res)
            return res
        
        if currency not in self.currencies.keys():
            res = ('Codigo de moneda base invalido: ' + currency)
            print(res)
            return res

        fecha_actual = datetime.now()
        mes = fecha_actual.month
        año = fecha_actual.year

        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

        res = self.get_month_exchange(mes, año, base_currency, currency)

        # Crear la gráfica
        plt.figure(figsize=(16, 10))
        plt.figure(figsize=(16, 10))
        fig = plt.figure()
        ax = fig.add_subplot(111)
        ax.set_facecolor('lavender')

        # Crear la gráfica
        plt.plot(meses[:mes], res, marker="D", markerfacecolor='lightpink', color='purple' ,linestyle='--', linewidth=2)  # Usamos 'marker' en vez de 'market'
        plt.grid(True, color='black', linestyle=':', linewidth=0.5)
        plt.title(f"Gráfica de tipo de cambio {currency}-{base_currency}")
        plt.xlabel("Meses")
        plt.ylabel("Valor")
        plt.xticks(rotation=45)

        # Guardamos la gráfica como archivo PNG
        nombre = "./graficas/historial.png"
        plt.savefig(nombre)

        print(f"Grafico creado en {nombre}")

    def graph_stats(self, base_currency, currency):

        if base_currency not in self.currencies.keys():
            res = ('Codigo de moneda base invalido: ' + base_currency)
            print(res)
            return res
        
        if currency not in self.currencies.keys():
            res = ('Codigo de moneda base invalido: ' + currency)
            print(res)
            return res

        fecha_actual = datetime.now()
        mes = fecha_actual.month
        año = fecha_actual.year

        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

        res = self.get_month_exchange(mes, año, base_currency, currency)

        dev_rate = []

        for i in range(len(res) - 1):
            initial_value = res[i]
            end_value = res[i + 1]
            dev_rate.append(devaluation_rate(initial_value, end_value))

        dev_rate_format = [round(n, 2) for n in dev_rate]
        
        # Crear la gráfica
        plt.figure(figsize=(16, 10))

        # Crear la gráfica
        fig = plt.figure()
        ax = fig.add_subplot(111)
        ax.set_facecolor('lavender')
        colormap = cm.coolwarm
        plt.bar(meses[:mes-1], dev_rate_format, color=colormap(range(len(meses))), edgecolor='black', linewidth=1.3, hatch='x')  # Usamos 'marker' en vez de 'market'
        plt.grid(True, color='black', linestyle='--', linewidth=0.5)
        plt.title(f"Devaluación mensual del {currency} frente al {base_currency}")
        plt.xlabel("Meses")
        plt.ylabel("Fluctuación")
        plt.xticks(rotation=45)

        # Guardamos la gráfica como archivo PNG
        nombre = "./graficas/stats.png"
        plt.savefig(nombre)

        print(f"Grafico creado en {nombre}")
